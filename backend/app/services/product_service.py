from app.utils.messages import Messages
from datetime import datetime, UTC
from io import BytesIO
import re

from fastapi import HTTPException, UploadFile
from pydantic import ValidationError
from app.database.mongodb import db
from app.utils.helpers import normalize_product_name, normalize_sku
from app.utils.responseBuilder import build_product_response
from app.utils.pagination import paginate_collection, regex_filter, validate_sort_field
from app.models.auth import UserRole
from app.models.product import ProductCreate
from app.services.supplier_service import get_own_company_supplier
from app.utils.settings import Settings
from app.core.exceptions import (
    forbidden,
    not_found,
    bad_request,
    conflict
)

products_collection = db.products
suppliers_collection = db.suppliers

PRODUCT_BULK_HEADERS = [
    "SKU",
    "Product Name",
    "Category",
    "Description",
    "Unit of Measure",
    "Tax Rate",
    "Reorder Level",
    "Supplier",
    "In-House",
    "Color",
    "Material",
    "Weight",
    "Size",
    "Dimension",
]

PRODUCT_BULK_HEADER_MAP = {
    "sku": "sku",
    "productname": "name",
    "category": "category",
    "description": "description",
    "unitofmeasure": "unit_of_measure",
    "taxrate": "tax_rate",
    "reorderlevel": "reorder_level",
    "supplier": "supplier_id",
    "inhouse": "is_manufactured",
    "color": "color",
    "material": "material",
    "weight": "weight",
    "size": "size",
    "dimension": "dimension",
}


def _normalize_header(value):
    return "".join(char for char in str(value or "").strip().lower() if char.isalnum())


def _clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_bool(value):
    text = _clean_cell(value).lower()
    if text in ["yes", "y", "true", "1", "in-house", "inhouse"]:
        return True
    if text in ["no", "n", "false", "0", ""]:
        return False
    raise ValueError("In-House must be Yes/No")


def _parse_float(value, field_name):
    text = _clean_cell(value)
    if text == "":
        return 0
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a number") from exc


def _parse_int(value, field_name):
    text = _clean_cell(value)
    if text == "":
        return 0
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a number") from exc


async def _resolve_supplier_id(value):
    supplier_value = _clean_cell(value)
    if not supplier_value:
        raise ValueError("Supplier is required")

    supplier = await suppliers_collection.find_one({
        "$or": [
            {"supplier_id": supplier_value},
            {"name": {"$regex": f"^{re.escape(supplier_value)}$", "$options": "i"}},
        ]
    })
    if not supplier:
        raise ValueError("Supplier not found")

    return supplier.get("supplier_id")


def _validation_reason(error):
    messages = []
    for item in error.errors():
        field = ".".join(str(part) for part in item.get("loc", []))
        messages.append(f"{field}: {item.get('msg')}")
    return "; ".join(messages) or "Invalid product data"


# CREATE PRODUCT
async def add_product(product_data: dict, auth_user: dict):
    sku = normalize_sku(product_data.get("sku"))
    product_data["name"] = normalize_product_name(product_data.get("name"))

    if auth_user.get("role") == UserRole.USER:
        forbidden()

    # Check duplicate SKU
    existing_product = await products_collection.find_one({
        "sku": sku
    })

    if existing_product:
        conflict(Messages.PRODUCT_ALREADY_EXISTS)

    own_supplier = await get_own_company_supplier()
    if product_data.get("is_manufactured") and own_supplier:
        product_data["supplier_id"] = own_supplier.get("supplier_id")

    supplier_id = product_data.get("supplier_id", "").strip()
    if not supplier_id:
        bad_request(Messages.INVALID_SUPPLIER_ID)
    product_data["supplier_id"] = supplier_id

    supplier = await suppliers_collection.find_one({
        "supplier_id": supplier_id
    })
    if not supplier:
        bad_request(Messages.INVALID_SUPPLIER_ID)

    product_data["created_at"] = datetime.now(UTC)
    product_data["updated_at"] = datetime.now(UTC)
    product_data["is_active"] = True
    product_data["is_manufactured"] = bool(supplier.get("is_own_company", False))
    product_data["sku"] = sku

    result = await products_collection.insert_one(product_data)

    created_product = await products_collection.find_one({
        "_id": result.inserted_id
    })

    return build_product_response(created_product)


async def bulk_upload_products(file: UploadFile, auth_user: dict):
    if auth_user.get("role") == UserRole.USER:
        forbidden()

    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        bad_request("Only .xlsx Excel files are allowed")

    contents = await file.read()
    max_size = Settings.BULK_UPLOAD_MAX_FILE_SIZE_MB * 1024 * 1024
    if len(contents) > max_size:
        bad_request(f"File size must be {Settings.BULK_UPLOAD_MAX_FILE_SIZE_MB}MB or less")

    try:
        from openpyxl import load_workbook
    except ImportError:
        bad_request("Excel upload support is not installed on the server")

    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        bad_request(f"Unable to read Excel file: {exc}")

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        bad_request("Excel file is empty")

    normalized_headers = [_normalize_header(header) for header in header_row]
    header_indexes = {}
    missing_headers = []

    for expected_header in PRODUCT_BULK_HEADERS:
        normalized = _normalize_header(expected_header)
        if normalized not in normalized_headers:
            missing_headers.append(expected_header)
        else:
            header_indexes[PRODUCT_BULK_HEADER_MAP[normalized]] = normalized_headers.index(normalized)

    if missing_headers:
        bad_request(f"Missing required columns: {', '.join(missing_headers)}")

    data_rows = [
        (row_number, row)
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2)
        if any(_clean_cell(value) for value in row)
    ]

    if not data_rows:
        bad_request("No product rows found in Excel file")

    if len(data_rows) > Settings.BULK_UPLOAD_MAX_ROWS:
        bad_request(f"Maximum {Settings.BULK_UPLOAD_MAX_ROWS} product rows allowed")

    row_results = []
    seen_skus = set()

    for row_number, row in data_rows:
        def cell(field):
            index = header_indexes[field]
            return row[index] if index < len(row) else None

        row_data = {}
        for header in PRODUCT_BULK_HEADERS:
            field = PRODUCT_BULK_HEADER_MAP[_normalize_header(header)]
            row_data[header] = _clean_cell(cell(field))

        try:
            supplier_id = await _resolve_supplier_id(cell("supplier_id"))
            sku = normalize_sku(_clean_cell(cell("sku")))
            if sku in seen_skus:
                raise ValueError("Duplicate SKU in uploaded file")
            seen_skus.add(sku)

            payload = {
                "sku": sku,
                "name": _clean_cell(cell("name")),
                "category": _clean_cell(cell("category")) or "General",
                "description": _clean_cell(cell("description")),
                "unit_of_measure": _clean_cell(cell("unit_of_measure")).lower(),
                "tax_rate": _parse_float(cell("tax_rate"), "Tax Rate"),
                "reorder_level": _parse_int(cell("reorder_level"), "Reorder Level"),
                "supplier_id": supplier_id,
                "is_manufactured": _parse_bool(cell("is_manufactured")),
                "attributes": {
                    "color": _clean_cell(cell("color")),
                    "material": _clean_cell(cell("material")),
                    "weight": _clean_cell(cell("weight")),
                    "size": _clean_cell(cell("size")),
                    "dimension": _clean_cell(cell("dimension")),
                },
            }
            validated_payload = ProductCreate(**payload).model_dump()
            created_product = await add_product(validated_payload, auth_user)

            row_results.append({
                "row_number": row_number,
                "status": "created",
                "reason": "",
                "data": row_data,
                "product": created_product,
            })
        except ValidationError as error:
            row_results.append({
                "row_number": row_number,
                "status": "failed",
                "reason": _validation_reason(error),
                "data": row_data,
            })
        except HTTPException as error:
            row_results.append({
                "row_number": row_number,
                "status": "failed",
                "reason": str(error.detail),
                "data": row_data,
            })
        except ValueError as error:
            row_results.append({
                "row_number": row_number,
                "status": "failed",
                "reason": str(error),
                "data": row_data,
            })

    created_count = len([row for row in row_results if row["status"] == "created"])
    failed_count = len([row for row in row_results if row["status"] == "failed"])
    return {
        "headers": PRODUCT_BULK_HEADERS,
        "summary": {
            "total": len(row_results),
            "created": created_count,
            "failed": failed_count,
        },
        "rows": row_results,
    }


# GET ALL PRODUCTS
async def get_all_products(
    auth_user: dict,
    search: str = None,
    sort_by: str = "created_at",
    order: str = "desc",
    page: int = 1,
    limit: int = 10
):
    allowed_sort_fields = [
        "created_at",
        "updated_at",
        "sku",
        "name",
        "category",
        "supplier_id",
        "tax_rate",
        "reorder_level",
        "is_active",
    ]
    validate_sort_field(sort_by, allowed_sort_fields)

    # superadmin can get all products, while admin and users can get active products
    if auth_user.get("role") == UserRole.SUPERADMIN:
        filters = {}

    elif auth_user.get("role") in [UserRole.ADMIN, UserRole.USER]:
        filters = {"is_active": True}

    else:
        forbidden()

    filters.update(regex_filter(search, ["sku", "name", "category", "supplier_id"]))

    return await paginate_collection(
        products_collection,
        filters,
        sort_by,
        order,
        page,
        limit,
        build_product_response,
    )


async def get_product_options(auth_user: dict, active_only: bool = False):
    filters = {}

    if active_only or auth_user.get("role") in [UserRole.ADMIN, UserRole.USER]:
        filters["is_active"] = True
    elif auth_user.get("role") != UserRole.SUPERADMIN:
        forbidden()

    products = []
    async for product in products_collection.find(
        filters,
        {"_id": 0, "sku": 1, "name": 1, "category": 1, "tax_rate": 1, "is_manufactured": 1}
    ).sort("name", 1):
        products.append({
            "sku": product.get("sku"),
            "name": product.get("name"),
            "category": product.get("category"),
            "tax_rate": product.get("tax_rate", 0),
            "is_manufactured": product.get("is_manufactured", False)
        })

    return products


async def get_product_form_options(auth_user: dict):
    product_filters = {}
    supplier_filters = {}

    if auth_user.get("role") in [UserRole.ADMIN, UserRole.USER]:
        product_filters["is_active"] = True
        supplier_filters["is_active"] = True
    elif auth_user.get("role") != UserRole.SUPERADMIN:
        forbidden()

    categories = await products_collection.distinct(
        "category",
        product_filters
    )

    suppliers = []
    async for supplier in suppliers_collection.find(
        supplier_filters,
            {"_id": 0, "supplier_id": 1, "name": 1, "is_own_company": 1}
    ).sort("name", 1):
        suppliers.append({
            "supplier_id": supplier.get("supplier_id"),
            "name": supplier.get("name"),
            "is_own_company": supplier.get("is_own_company", False),
        })

    return {
        "categories": sorted([category for category in categories if category]),
        "suppliers": suppliers,
        "units": ["pcs", "kg", "g", "m", "cm", "ltr", "ml", "other"]
    }


# GET SINGLE PRODUCT
async def get_product_by_sku(sku: str, auth_user: dict):
    sku = normalize_sku(sku)

    if not sku:
        bad_request(Messages.INVALID_SKU)

    product = await products_collection.find_one({
        "sku": sku
    })

    if not product:
        not_found(Messages.PRODUCT_NOT_FOUND)

    if auth_user.get("role") == UserRole.SUPERADMIN:
        return build_product_response(product)

    if auth_user.get("role") in [UserRole.ADMIN, UserRole.USER]:

        if not product.get("is_active"):
            forbidden()
        return build_product_response(product)

    forbidden()


# UPDATE PRODUCT
async def update_product_by_sku(sku: str, update_data: dict, auth_user: dict):

    if auth_user.get("role") == UserRole.USER:
        forbidden()

    sku = normalize_sku(sku)

    if not sku:
        bad_request(Messages.INVALID_SKU)

    existing_product = await products_collection.find_one({
        "sku": sku
    })

    if not existing_product:
        not_found(Messages.PRODUCT_NOT_FOUND)

    if (
        auth_user.get("role") == UserRole.ADMIN
        and not existing_product.get("is_active")
    ):
        forbidden()

    if "name" in update_data:
        update_data["name"] = normalize_product_name(update_data.get("name"))

    own_supplier = await get_own_company_supplier()
    if update_data.get("is_manufactured") and own_supplier:
        update_data["supplier_id"] = own_supplier.get("supplier_id")

    supplier_id = update_data.get("supplier_id")
    if supplier_id is not None:
        supplier = await suppliers_collection.find_one({
            "supplier_id": supplier_id
        })
        if not supplier:
            bad_request(Messages.INVALID_SUPPLIER_ID)
        update_data["is_manufactured"] = bool(supplier.get("is_own_company", False))

    update_data["updated_at"] = datetime.now(UTC)

    await products_collection.update_one(
        {"sku": sku},
        {"$set": update_data}
    )

    updated_product = await products_collection.find_one({
        "sku": sku
    })

    return build_product_response(updated_product)


# DELETE PRODUCT
async def delete_product_by_sku(sku: str, auth_user: dict):
    if auth_user.get("role") == UserRole.USER:
        forbidden()

    sku = normalize_sku(sku)

    if not sku:
        bad_request(Messages.INVALID_SKU)

    existing_product = await products_collection.find_one({"sku": sku})

    if not existing_product:
        not_found(Messages.PRODUCT_NOT_FOUND)

    if auth_user.get("role") == UserRole.SUPERADMIN:
        if not existing_product.get("is_active"):
            await products_collection.delete_one({"sku": sku})
            return Messages.PRODUCT_DELETED_PERMANENTLY

        await products_collection.update_one(
            {"sku": sku},
            {"$set": {
                "is_active": False,
                "updated_at": datetime.now(UTC)
            }}
        )
        return Messages.PRODUCT_DELETED

    if auth_user.get("role") == UserRole.ADMIN:
        if not existing_product.get("is_active"):
            bad_request(Messages.PRODUCT_INACTIVE)

        await products_collection.update_one(
            {"sku": sku},
            {"$set": {
                "is_active": False,
                "updated_at": datetime.now(UTC)
            }}
        )
        return Messages.PRODUCT_DELETED

    forbidden()
 

async def filter_products_service(
    sku=None, name=None, category=None,
    supplier_id=None, is_active=None, auth_user=None
):
    filters = {}
    products = []

    if auth_user.get("role") != UserRole.SUPERADMIN:
        if is_active is None:
            filters["is_active"] = True

        elif not is_active:
            return products

        filters["is_active"] = True

    else:
        if is_active is not None:
            filters["is_active"] = is_active

    if sku:
        filters["sku"] = normalize_sku(sku)

    if name:
        filters["name"] = normalize_product_name(name)

    if category:
        filters["category"] = category

    if supplier_id:
        filters["supplier_id"] = supplier_id

    async for product in products_collection.find(filters).sort("created_at", -1):
        products.append(
            build_product_response(product)
        )

    return products
