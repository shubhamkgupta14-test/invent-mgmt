from io import BytesIO
from zipfile import BadZipFile, ZipFile

from fastapi import UploadFile

from app.core.exceptions import bad_request
from app.utils.settings import Settings


def normalize_header(value):
    return "".join(char for char in str(value or "").strip().lower() if char.isalnum())


def clean_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_float(value, field_name, default=0):
    text = clean_cell(value)
    if text == "":
        return default
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a number") from exc


def parse_int(value, field_name, default=0):
    text = clean_cell(value)
    if text == "":
        return default
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(f"{field_name} must be a number") from exc


async def read_bulk_excel(file: UploadFile, headers: list[str], header_map: dict[str, str]):
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        bad_request("Only .xlsx Excel files are allowed")

    contents = await file.read()
    max_size = Settings.BULK_UPLOAD_MAX_FILE_SIZE_MB * 1024 * 1024
    if len(contents) > max_size:
        bad_request(f"File size must be {Settings.BULK_UPLOAD_MAX_FILE_SIZE_MB}MB or less")

    try:
        with ZipFile(BytesIO(contents)) as archive:
            entries = archive.infolist()
            total_uncompressed = sum(entry.file_size for entry in entries)
            if len(entries) > 200:
                bad_request("Excel file contains too many internal entries")
            if total_uncompressed > Settings.BULK_UPLOAD_MAX_UNCOMPRESSED_MB * 1024 * 1024:
                bad_request("Excel file expands beyond the allowed size")
            for entry in entries:
                if entry.flag_bits & 0x1:
                    bad_request("Encrypted Excel files are not allowed")
                if entry.file_size > 10 * 1024 * 1024:
                    bad_request("An Excel file entry is too large")
                if entry.compress_size and entry.file_size / entry.compress_size > 200:
                    bad_request("Excel file compression ratio is unsafe")
    except BadZipFile:
        bad_request("Invalid .xlsx Excel file")

    try:
        from openpyxl import load_workbook
    except ImportError:
        bad_request("Excel upload support is not installed on the server")

    try:
        workbook = load_workbook(BytesIO(contents), read_only=True, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        bad_request(f"Unable to read Excel file: {exc}")

    if sheet.max_row > Settings.BULK_UPLOAD_MAX_ROWS + 1:
        bad_request(f"Maximum {Settings.BULK_UPLOAD_MAX_ROWS} rows allowed")
    if sheet.max_column > 100:
        bad_request("Excel file contains too many columns")

    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        bad_request("Excel file is empty")

    normalized_headers = [normalize_header(header) for header in header_row]
    header_indexes = {}
    missing_headers = []

    for expected_header in headers:
        normalized = normalize_header(expected_header)
        if normalized not in normalized_headers:
            missing_headers.append(expected_header)
        else:
            header_indexes[header_map[normalized]] = normalized_headers.index(normalized)

    if missing_headers:
        bad_request(f"Missing required columns: {', '.join(missing_headers)}")

    data_rows = [
        (row_number, row)
        for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2)
        if any(clean_cell(value) for value in row)
    ]

    if not data_rows:
        bad_request("No rows found in Excel file")

    if len(data_rows) > Settings.BULK_UPLOAD_MAX_ROWS:
        bad_request(f"Maximum {Settings.BULK_UPLOAD_MAX_ROWS} rows allowed")

    return header_indexes, data_rows


def get_row_cell(row, header_indexes: dict[str, int], field: str):
    index = header_indexes[field]
    return row[index] if index < len(row) else None


def build_row_data(headers: list[str], header_map: dict[str, str], header_indexes: dict[str, int], row):
    data = {}
    for header in headers:
        field = header_map[normalize_header(header)]
        data[header] = clean_cell(get_row_cell(row, header_indexes, field))
    return data


def summarize_bulk_rows(headers: list[str], rows: list[dict]):
    created_count = len([row for row in rows if row["status"] == "created"])
    failed_count = len([row for row in rows if row["status"] == "failed"])
    return {
        "headers": headers,
        "summary": {
            "total": len(rows),
            "created": created_count,
            "failed": failed_count,
        },
        "rows": rows,
    }
